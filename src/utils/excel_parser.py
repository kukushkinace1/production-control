from __future__ import annotations

import csv
from datetime import date, datetime, time
from io import BytesIO, StringIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

IMPORT_COLUMNS = {
    "batch_number": ("batch_number", "НомерПартии"),
    "batch_date": ("batch_date", "ДатаПартии"),
    "task_description": ("task_description", "ПредставлениеЗаданияНаСмену"),
    "work_center_name": ("work_center_name", "РабочийЦентр"),
    "work_center_identifier": ("work_center_identifier", "ИдентификаторРЦ"),
    "shift": ("shift", "Смена"),
    "team": ("team", "Бригада"),
    "nomenclature": ("nomenclature", "Номенклатура"),
    "ekn_code": ("ekn_code", "КодЕКН"),
    "shift_start": ("shift_start", "ДатаВремяНачалаСмены"),
    "shift_end": ("shift_end", "ДатаВремяОкончанияСмены"),
    "is_closed": ("is_closed", "СтатусЗакрытия"),
}

EXPORT_HEADERS = [
    "id",
    "batch_number",
    "batch_date",
    "is_closed",
    "work_center_identifier",
    "work_center_name",
    "shift",
    "team",
    "nomenclature",
    "ekn_code",
    "task_description",
    "shift_start",
    "shift_end",
    "products_total",
    "products_aggregated",
]


def parse_batch_import_file(
    data: bytes,
    file_name: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    suffix = file_name.lower().rsplit(".", 1)[-1] if "." in file_name else ""
    if suffix == "csv":
        raw_rows = _read_csv_rows(data)
    elif suffix in {"xlsx", "xlsm"}:
        raw_rows = _read_excel_rows(data)
    else:
        return [], [{"row": 0, "error": "Unsupported file format. Use .csv or .xlsx."}]

    rows: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    seen_keys: set[tuple[int, date]] = set()

    for row_number, raw_row in raw_rows:
        try:
            row = validate_batch_import_row(raw_row)
        except ValueError as exc:
            errors.append({"row": row_number, "error": str(exc)})
            continue

        key = (row["batch_number"], row["batch_date"])
        if key in seen_keys:
            errors.append({"row": row_number, "error": "Duplicate row in import file."})
            continue

        seen_keys.add(key)
        row["_row_number"] = row_number
        rows.append(row)

    return rows, errors


def validate_batch_import_row(raw_row: dict[str, Any]) -> dict[str, Any]:
    normalized = {_normalize_header(key): value for key, value in raw_row.items()}
    row: dict[str, Any] = {}

    for field, aliases in IMPORT_COLUMNS.items():
        value = _get_by_alias(normalized, aliases)
        if field == "is_closed" and value in (None, ""):
            row[field] = False
            continue
        if value in (None, ""):
            raise ValueError(f"Missing required field: {field}")
        row[field] = value

    row["batch_number"] = _parse_int(row["batch_number"], "batch_number")
    row["batch_date"] = _parse_date(row["batch_date"], "batch_date")
    row["shift_start"] = _parse_datetime(row["shift_start"], "shift_start")
    row["shift_end"] = _parse_datetime(row["shift_end"], "shift_end")
    row["is_closed"] = _parse_bool(row["is_closed"])

    for field in (
        "task_description",
        "work_center_name",
        "work_center_identifier",
        "shift",
        "team",
        "nomenclature",
        "ekn_code",
    ):
        row[field] = str(row[field]).strip()
        if not row[field]:
            raise ValueError(f"Missing required field: {field}")

    if row["shift_end"] <= row["shift_start"]:
        raise ValueError("shift_end must be after shift_start")

    return row


def generate_batches_export_excel(rows: list[dict[str, Any]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "batches"
    sheet.append(EXPORT_HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        sheet.append([_stringify_export_value(row.get(header)) for header in EXPORT_HEADERS])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def generate_batches_export_csv(rows: list[dict[str, Any]]) -> bytes:
    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=EXPORT_HEADERS, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow(
            {header: _stringify_export_value(row.get(header)) for header in EXPORT_HEADERS}
        )
    return output.getvalue().encode("utf-8-sig")


def _read_csv_rows(data: bytes) -> list[tuple[int, dict[str, Any]]]:
    text = data.decode("utf-8-sig")
    reader = csv.DictReader(StringIO(text))
    return [(index, dict(row)) for index, row in enumerate(reader, start=2)]


def _read_excel_rows(data: bytes) -> list[tuple[int, dict[str, Any]]]:
    workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    try:
        headers = next(iterator)
    except StopIteration:
        return []

    rows: list[tuple[int, dict[str, Any]]] = []
    for row_number, values in enumerate(iterator, start=2):
        if all(value in (None, "") for value in values):
            continue
        rows.append((row_number, dict(zip(headers, values, strict=False))))
    return rows


def _get_by_alias(row: dict[str, Any], aliases: tuple[str, ...]) -> Any:
    for alias in aliases:
        value = row.get(_normalize_header(alias))
        if value not in (None, ""):
            return value
    return None


def _normalize_header(header: Any) -> str:
    return str(header or "").strip().lower().replace(" ", "").replace("_", "")


def _parse_int(value: Any, field_name: str) -> int:
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Invalid integer field: {field_name}") from exc


def _parse_date(value: Any, field_name: str) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Invalid date field: {field_name}")


def _parse_datetime(value: Any, field_name: str) -> datetime:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    text = str(value).strip().replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(text)
    except ValueError as exc:
        raise ValueError(f"Invalid datetime field: {field_name}") from exc


def _parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return bool(value)
    text = str(value).strip().lower()
    return text in {"true", "1", "yes", "y", "да", "закрыта", "closed"}


def _stringify_export_value(value: Any) -> Any:
    if isinstance(value, datetime | date):
        return value.isoformat()
    return value
