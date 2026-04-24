from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font


def generate_batch_report_excel(report_data: dict) -> bytes:
    workbook = Workbook()

    _build_info_sheet(workbook.active, report_data)
    _build_products_sheet(workbook.create_sheet("products"), report_data)
    _build_statistics_sheet(workbook.create_sheet("statistics"), report_data)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _build_info_sheet(sheet, report_data: dict) -> None:
    sheet.title = "info"
    sheet.append(["Field", "Value"])
    batch = report_data["batch"]
    rows = [
        ("Batch ID", batch["id"]),
        ("Batch Number", batch["batch_number"]),
        ("Batch Date", batch["batch_date"]),
        ("Task Description", batch["task_description"]),
        ("Work Center", batch["work_center_name"]),
        ("Work Center Identifier", batch["work_center_identifier"]),
        ("Shift", batch["shift"]),
        ("Team", batch["team"]),
        ("Nomenclature", batch["nomenclature"]),
        ("EKN Code", batch["ekn_code"]),
        ("Closed", batch["is_closed"]),
        ("Shift Start", batch["shift_start"]),
        ("Shift End", batch["shift_end"]),
        ("Generated At", report_data["generated_at"]),
    ]
    for row in rows:
        sheet.append(list(row))
    _format_header(sheet)


def _build_products_sheet(sheet, report_data: dict) -> None:
    sheet.append(["Unique Code", "Aggregated", "Aggregated At", "Created At"])
    for product in report_data["products"]:
        sheet.append(
            [
                product["unique_code"],
                product["is_aggregated"],
                product["aggregated_at"],
                product["created_at"],
            ]
        )
    _format_header(sheet)


def _build_statistics_sheet(sheet, report_data: dict) -> None:
    sheet.append(["Metric", "Value"])
    statistics = report_data["statistics"]
    rows = [
        ("Total Products", statistics["total_products"]),
        ("Aggregated Products", statistics["aggregated_products"]),
        ("Remaining Products", statistics["remaining_products"]),
        ("Aggregation Rate", statistics["aggregation_rate"]),
        ("Shift Duration Hours", statistics["shift_duration_hours"]),
    ]
    for row in rows:
        sheet.append(list(row))
    _format_header(sheet)


def _format_header(sheet) -> None:
    for cell in sheet[1]:
        cell.font = Font(bold=True)
