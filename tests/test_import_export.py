from datetime import datetime
from io import BytesIO
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from src.utils import (
    generate_batches_export_excel,
    parse_batch_import_file,
    validate_batch_import_row,
)


def build_import_row() -> dict:
    return {
        "batch_number": "3001",
        "batch_date": "2026-04-24",
        "task_description": "Day shift",
        "work_center_name": "Line 1",
        "work_center_identifier": "WC-1",
        "shift": "day",
        "team": "A",
        "nomenclature": "Widget",
        "ekn_code": "EKN-1",
        "shift_start": "2026-04-24T08:00:00+00:00",
        "shift_end": "2026-04-24T20:00:00+00:00",
        "is_closed": "false",
    }


def test_validate_batch_import_row_normalizes_values() -> None:
    row = validate_batch_import_row(build_import_row())

    assert row["batch_number"] == 3001
    assert row["batch_date"].isoformat() == "2026-04-24"
    assert row["shift_start"] == datetime.fromisoformat("2026-04-24T08:00:00+00:00")
    assert row["is_closed"] is False


def test_parse_batch_import_file_reports_duplicate_rows() -> None:
    content = (
        b"batch_number,batch_date,task_description,work_center_name,work_center_identifier,"
        b"shift,team,nomenclature,ekn_code,shift_start,shift_end,is_closed\n"
        b"3001,2026-04-24,Day shift,Line 1,WC-1,day,A,Widget,EKN-1,"
        b"2026-04-24T08:00:00+00:00,2026-04-24T20:00:00+00:00,false\n"
        b"3001,2026-04-24,Day shift,Line 1,WC-1,day,A,Widget,EKN-1,"
        b"2026-04-24T08:00:00+00:00,2026-04-24T20:00:00+00:00,false\n"
    )

    rows, errors = parse_batch_import_file(content, "batches.csv")

    assert len(rows) == 1
    assert errors == [{"row": 3, "error": "Duplicate row in import file."}]


def test_generate_batches_export_excel_smoke() -> None:
    workbook_bytes = generate_batches_export_excel(
        [
            {
                "id": 1,
                "batch_number": 3001,
                "batch_date": "2026-04-24",
                "is_closed": False,
                "work_center_identifier": "WC-1",
                "work_center_name": "Line 1",
                "shift": "day",
                "team": "A",
                "nomenclature": "Widget",
                "ekn_code": "EKN-1",
                "task_description": "Day shift",
                "shift_start": "2026-04-24T08:00:00+00:00",
                "shift_end": "2026-04-24T20:00:00+00:00",
                "products_total": 10,
                "products_aggregated": 7,
            }
        ]
    )

    workbook = load_workbook(BytesIO(workbook_bytes))

    assert workbook.sheetnames == ["batches"]
    assert workbook["batches"]["A1"].value == "id"
    assert workbook["batches"]["B2"].value == 3001


@pytest.mark.asyncio
async def test_import_batches_endpoint_uploads_file_and_starts_task(client, monkeypatch) -> None:
    from src.api.v1.routers import batches as batches_router_module

    class FakeMinIOService:
        def ensure_bucket(self, bucket: str) -> None:
            assert bucket == "imports"

        def upload_bytes(self, **kwargs) -> str:
            assert kwargs["bucket"] == "imports"
            assert kwargs["data"] == b"file-content"
            return "http://localhost:9000/imports/file.csv"

    monkeypatch.setattr(batches_router_module, "MinIOService", FakeMinIOService)
    monkeypatch.setattr(
        batches_router_module.import_batches_from_file,
        "delay",
        lambda **_: SimpleNamespace(id="import-task-123"),
    )

    response = await client.post(
        "/api/v1/batches/import",
        files={"file": ("batches.csv", b"file-content", "text/csv")},
    )

    assert response.status_code == 202
    assert response.json() == {
        "task_id": "import-task-123",
        "status": "PENDING",
        "message": "File uploaded, import started",
    }


@pytest.mark.asyncio
async def test_export_batches_endpoint_starts_task(client, monkeypatch) -> None:
    from src.api.v1.routers import batches as batches_router_module

    monkeypatch.setattr(
        batches_router_module.export_batches_to_file,
        "delay",
        lambda **_: SimpleNamespace(id="export-task-123"),
    )

    response = await client.post(
        "/api/v1/batches/export",
        json={
            "format": "excel",
            "filters": {
                "is_closed": False,
                "date_from": "2026-04-01",
                "date_to": "2026-04-30",
            },
        },
    )

    assert response.status_code == 202
    assert response.json() == {
        "task_id": "export-task-123",
        "status": "PENDING",
        "message": "Batch export task started",
    }
