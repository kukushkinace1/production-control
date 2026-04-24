from io import BytesIO
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from src.main import app
from src.utils import generate_batch_report_excel, generate_batch_report_pdf


def build_report_data() -> dict:
    return {
        "batch": {
            "id": 1,
            "batch_number": 101,
            "batch_date": "2026-04-20",
            "task_description": "Night shift batch",
            "work_center_name": "Line 1",
            "work_center_identifier": "WC-1",
            "shift": "night",
            "team": "A",
            "nomenclature": "Widget",
            "ekn_code": "EKN-101",
            "is_closed": False,
            "shift_start": "2026-04-20T08:00:00Z",
            "shift_end": "2026-04-20T20:00:00Z",
        },
        "products": [
            {
                "unique_code": "CODE-001",
                "is_aggregated": True,
                "aggregated_at": "2026-04-20T09:00:00Z",
                "created_at": "2026-04-20T08:10:00Z",
            },
            {
                "unique_code": "CODE-002",
                "is_aggregated": False,
                "aggregated_at": None,
                "created_at": "2026-04-20T08:11:00Z",
            },
        ],
        "statistics": {
            "total_products": 2,
            "aggregated_products": 1,
            "remaining_products": 1,
            "aggregation_rate": 50.0,
            "shift_duration_hours": 12.0,
        },
        "generated_at": "2026-04-20T10:00:00Z",
    }


def test_generate_batch_report_excel_creates_expected_sheets() -> None:
    workbook_bytes = generate_batch_report_excel(build_report_data())
    workbook = load_workbook(BytesIO(workbook_bytes))

    assert workbook.sheetnames == ["info", "products", "statistics"]
    assert workbook["info"]["A2"].value == "Batch ID"
    assert workbook["info"]["B2"].value == 1
    assert workbook["products"]["A2"].value == "CODE-001"
    assert workbook["statistics"]["A2"].value == "Total Products"
    assert workbook["statistics"]["B2"].value == 2


def test_generate_batch_report_pdf_creates_pdf_bytes() -> None:
    pdf_bytes = generate_batch_report_pdf(build_report_data())

    assert pdf_bytes.startswith(b"%PDF-1.4")
    assert b"%%EOF" in pdf_bytes


@pytest.mark.asyncio
async def test_create_batch_report_returns_task_id(client, monkeypatch) -> None:
    from src.api.v1.routers import batches as batches_router_module
    from src.core.dependencies import get_batch_service

    class FakeBatchService:
        async def get_batch(self, batch_id: int):
            return SimpleNamespace(id=batch_id)

    app.dependency_overrides[get_batch_service] = lambda: FakeBatchService()
    monkeypatch.setattr(
        batches_router_module.generate_batch_report,
        "delay",
        lambda **_: SimpleNamespace(id="report-task-123"),
    )

    response = await client.post(
        "/api/v1/batches/1/reports",
        json={"format": "excel"},
    )

    assert response.status_code == 202
    assert response.json() == {
        "task_id": "report-task-123",
        "status": "PENDING",
        "message": "Report generation task started",
    }
